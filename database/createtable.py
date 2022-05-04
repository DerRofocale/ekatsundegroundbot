from database import GetAllAdmins, GetAllStats
from database import GetAllUsers
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font


async def CreateTableUsers():
    book = openpyxl.Workbook() # СОЗДАЁМ ФАЙЛ
    book.remove(book.active) # УДАЛЯЕМ СТАНДАРТНУЮ СТРАНИЦУ SHEET
    sheet_1 = book.create_sheet("Пользователи") # СОЗДАЁМ ЛИСТ ПОЛЬЗОВАТЕЛИ
    standart = ["Телеграм ID", "Полное имя пользователя", "Ссылка пользователя", "Дата и время регистрации в боте"]  # УКАЗЫВАЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    sheet_1.append(standart)  # ДОБАВЛЯЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    for usr in await GetAllUsers():  # ДОБАВЛЯЕМ МАССИВЫ ДАННЫХ
        sheet_1.append(usr)
    sheet_1.auto_filter.ref = "A:D"  # ДОБАВЛЯЕМ АВТОФИЛЬТРЫ СТОЛБЦАМ
    # УКАЗЫВАЕМ РАЗМЕР СТОЛБЦАМ
    sheet_1.column_dimensions["A"].width = 15
    sheet_1.column_dimensions["B"].width = 35
    sheet_1.column_dimensions["C"].width = 35
    sheet_1.column_dimensions["D"].width = 35


    # СТИЛИЗУЕМ ЗАГОЛОВКИ
    font = Font(b=True)
    sheet_1["A1"].font = font
    sheet_1["B1"].font = font
    sheet_1["C1"].font = font
    sheet_1["D1"].font = font

    date = datetime.now().strftime("%H-%M-%S %d.%m.%Y")

    bio = BytesIO()
    book.save(bio)
    bio.name = f"Пользователи {date}.xlsx"
    bio.seek(0)
    return bio


async def CreateTableAdmins():
    book = openpyxl.Workbook() # СОЗДАЁМ ФАЙЛ
    book.remove(book.active) # УДАЛЯЕМ СТАНДАРТНУЮ СТРАНИЦУ SHEET
    sheet_1 = book.create_sheet("Администраторы") # СОЗДАЁМ ЛИСТ ПОЛЬЗОВАТЕЛИ
    standart = ["Телеграм ID", "Полное имя администратора", "Ссылка администратора", "Дата и время назначения прав"]  # УКАЗЫВАЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    sheet_1.append(standart)  # ДОБАВЛЯЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    for usr in await GetAllAdmins():  # ДОБАВЛЯЕМ МАССИВЫ ДАННЫХ
        sheet_1.append(usr)
    sheet_1.auto_filter.ref = "A:D"  # ДОБАВЛЯЕМ АВТОФИЛЬТРЫ СТОЛБЦАМ
    # УКАЗЫВАЕМ РАЗМЕР СТОЛБЦАМ
    sheet_1.column_dimensions["A"].width = 15
    sheet_1.column_dimensions["B"].width = 35
    sheet_1.column_dimensions["C"].width = 35
    sheet_1.column_dimensions["D"].width = 35


    # СТИЛИЗУЕМ ЗАГОЛОВКИ
    font = Font(b=True)
    sheet_1["A1"].font = font
    sheet_1["B1"].font = font
    sheet_1["C1"].font = font
    sheet_1["D1"].font = font

    date = datetime.now().strftime("%H-%M-%S %d.%m.%Y")

    bio = BytesIO()
    book.save(bio)
    bio.name = f"Администраторы {date}.xlsx"
    bio.seek(0)
    return bio

async def CreateTableStats():
    book = openpyxl.Workbook() # СОЗДАЁМ ФАЙЛ
    book.remove(book.active) # УДАЛЯЕМ СТАНДАРТНУЮ СТРАНИЦУ SHEET
    sheet_1 = book.create_sheet("Статистика") # СОЗДАЁМ ЛИСТ ПОЛЬЗОВАТЕЛИ
    standart = ["Телеграм ID", "Станция отправления", "Направление движения", "Дата и время запроса"]  # УКАЗЫВАЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    sheet_1.append(standart)  # ДОБАВЛЯЕМ ЗАГОЛОВКИ СТОЛБЦОВ
    for usr in await GetAllStats():  # ДОБАВЛЯЕМ МАССИВЫ ДАННЫХ
        sheet_1.append(usr)
    sheet_1.auto_filter.ref = "A:D"  # ДОБАВЛЯЕМ АВТОФИЛЬТРЫ СТОЛБЦАМ
    # УКАЗЫВАЕМ РАЗМЕР СТОЛБЦАМ
    sheet_1.column_dimensions["A"].width = 15
    sheet_1.column_dimensions["B"].width = 25
    sheet_1.column_dimensions["C"].width = 25
    sheet_1.column_dimensions["D"].width = 25


    # СТИЛИЗУЕМ ЗАГОЛОВКИ
    font = Font(b=True)
    sheet_1["A1"].font = font
    sheet_1["B1"].font = font
    sheet_1["C1"].font = font
    sheet_1["D1"].font = font

    date = datetime.now().strftime("%H-%M-%S %d.%m.%Y")

    bio = BytesIO()
    book.save(bio)
    bio.name = f"Статистика {date}.xlsx"
    bio.seek(0)
    return bio

